from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('CRA_signup.xlsx')
ws = wb.active

name:str = ""
nomul:int = 0
sat1:int = 0
sat2:int = 0
sun1:int = 0
sun2:int = 0
mon:int = 0
count:int = 0
index1:int = 3
index2:int = 4

for fuckitweball in range(34):
    for row in range(index1,index2):
        for col in range(8,14):
            char = get_column_letter(col)
            #print(ws[char + str(row)].value)
            if col == 8:
                name = ws[char + str(row)].value 
            if col == 9:
                if ws[char + str(row)].value != None:
                    sat1 = sat1+1
            if col == 10:
                if ws[char + str(row)].value != None:
                    sat2 = sat2+1
            if col == 11:
                if ws[char + str(row)].value != None:
                    sun1 = sun1+1        
            if col == 12:
                if ws[char + str(row)].value != None:
                    sun2 = sun2+1    
            if col == 13:
                if ws[char + str(row)].value != None:
                    mon = mon+1 

    if sat1 == 1:
        for row in range(3,13):
            for col in range(2,3):
                char = get_column_letter(col)
                if ws[char + str(row)].value == None and count == 0 and nomul == 0:
                    ws[char + str(row)] = name 
                    count = count+1
                    nomul = nomul+1
    count = 0
    sat1 = 0

    if sat2 == 1:
        for row in range(3,13):
            for col in range(3,4):
                char = get_column_letter(col)
                if ws[char + str(row)].value == None and count == 0 and nomul == 0:
                    ws[char + str(row)] = name 
                    count = count+1
                    nomul = nomul+1
    count = 0
    sat2 = 0

    if sun1 == 1:
        for row in range(3,13):
            for col in range(4,5):
                char = get_column_letter(col)
                if ws[char + str(row)].value == None and count == 0 and nomul == 0:
                    ws[char + str(row)] = name 
                    count = count+1
                    nomul = nomul+1
    count = 0
    sun1 = 0

    if sun2 == 1:
        for row in range(3,13):
            for col in range(5,6):
                char = get_column_letter(col)
                if ws[char + str(row)].value == None and count == 0 and nomul == 0:
                    ws[char + str(row)] = name 
                    count = count+1
                    nomul = nomul+1
    count = 0
    sun2 = 0

    if mon == 1:
        for row in range(3,13):
            for col in range(6,7):
                char = get_column_letter(col)
                if ws[char + str(row)].value == None and count == 0 and nomul == 0:
                    ws[char + str(row)] = name 
                    count = count+1
                    nomul = nomul+1
    count = 0
    mon = 0

    nomul = 0

    index1=index1+1
    index2=index2+1

wb.save('CRA_signup.xlsx')


